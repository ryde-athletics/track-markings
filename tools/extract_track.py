#!/usr/bin/env python3
"""One-time extractor: DunbarLayout.xlsx -> track.json for the web field map.

The workbook holds fixed planar geometry (metres, local grid; origin at mark "X",
the home straight runs along +X). This pulls the survey marks and the drawn track
context out of the sheets that feed the "Plan" tab's scatter chart and writes a
compact JSON blob that gets pasted into index.html as `const TRACK = {...}`.

Run from the repo root:  python3 tools/extract_track.py
Then copy track.json's contents into the TRACK const in index.html.
"""

import json
import math
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "DunbarLayout.xlsx"
OUT = ROOT / "track.json"

R = 3  # decimal places


def rnd(v):
    return round(float(v), R)


def col_pairs(ws, xcol, ycol, r0, r1):
    """Collect [x, y] rows from two columns, skipping blanks."""
    out = []
    for row in range(r0, r1 + 1):
        x = ws[f"{xcol}{row}"].value
        y = ws[f"{ycol}{row}"].value
        if x is None or y is None:
            continue
        try:
            out.append([rnd(x), rnd(y)])
        except (TypeError, ValueError):
            continue
    return out


def main():
    if not XLSX.exists():
        sys.exit(f"not found: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # --- survey marks: All!A category, B label, C x, D y --------------------
    all_ws = wb["All"]
    marks = []
    seen = set()
    for row in range(2, all_ws.max_row + 1):
        label = all_ws[f"B{row}"].value
        x = all_ws[f"C{row}"].value
        y = all_ws[f"D{row}"].value
        cat = all_ws[f"A{row}"].value
        if label is None or x is None or y is None:
            continue
        label = str(label).strip()
        if not label or label in seen:
            continue
        seen.add(label)
        cat = str(cat).strip() if cat else "Other"
        if cat == "Pack Start":  # pack starts are just another kind of start line
            cat = "Starts"
        marks.append({
            "id": len(marks),
            "label": label,
            "cat": cat,
            "x": rnd(x),
            "y": rnd(y),
        })

    # --- 9 lane "line of running" polylines: Lanes rows 3..77 --------------
    lanes_ws = wb["Lanes"]
    lane_cols = [("D", "E"), ("I", "J"), ("N", "O"), ("S", "T"), ("X", "Y"),
                 ("AC", "AD"), ("AH", "AI"), ("AM", "AN"), ("AR", "AS")]
    lanes = [col_pairs(lanes_ws, xc, yc, 3, 77) for xc, yc in lane_cols]

    # --- straight-line features from Lanes --------------------------------
    lines = {
        "finish": col_pairs(lanes_ws, "AV", "AW", 3, 4),
        "homeStraight": [],
    }
    # Home-straight series live in AX/AY as 4-row blocks: label, "X"/"Y" header,
    # then two point rows. Blocks start at rows 1, 5, 9, ... 37.
    hs = []
    for base in range(1, 38, 4):
        seg = col_pairs(lanes_ws, "AX", "AY", base + 2, base + 3)
        if len(seg) == 2:
            hs.append(seg)
    lines["homeStraight"] = hs

    # --- context point clouds (non-interactive dots) ---------------------
    hurdles = col_pairs(wb["Hurdles"], "F", "G", 3, 82)
    relays_ws = wb["Relays"]
    relays = (col_pairs(relays_ws, "D", "E", 3, 26)
              + col_pairs(relays_ws, "I", "J", 3, 26)
              + col_pairs(relays_ws, "N", "O", 3, 26))

    track = {
        "units": "metres",
        "marks": marks,
        "lanes": lanes,
        "lines": lines,
        "hurdles": hurdles,
        "relays": relays,
    }

    OUT.write_text(json.dumps(track, separators=(",", ":")))
    print(f"wrote {OUT}  ({OUT.stat().st_size} bytes)")
    cats = {}
    for m in marks:
        cats[m["cat"]] = cats.get(m["cat"], 0) + 1
    print(f"marks: {len(marks)}  by category: {cats}")
    print(f"lanes: {[len(l) for l in lanes]}")
    print(f"context: hurdles={len(hurdles)} relays={len(relays)} "
          f"finish={len(lines['finish'])} homeStraight={len(lines['homeStraight'])}")

    # --- sanity checks against the workbook's own numbers ----------------
    by_label = {m["label"]: m for m in marks}

    def dist(a, b):
        pa, pb = by_label[a], by_label[b]
        return math.hypot(pb["x"] - pa["x"], pb["y"] - pa["y"])

    print("\nsanity (expect ~76.89 m each):")
    ok = True
    for a, b in [("C", "D"), ("A", "B"), ("X", "Y")]:
        d = dist(a, b)
        flag = "ok" if abs(d - 76.89) < 0.05 else "MISMATCH"
        ok &= flag == "ok"
        print(f"  {a} -> {b}: {d:.2f}  {flag}")
    xs = [m["x"] for m in marks]
    ys = [m["y"] for m in marks]
    print(f"\nbounds: x [{min(xs):.1f}, {max(xs):.1f}]  y [{min(ys):.1f}, {max(ys):.1f}]")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
